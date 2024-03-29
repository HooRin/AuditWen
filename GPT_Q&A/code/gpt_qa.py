import os

import openpyxl
import pandas as pd
from openai import OpenAI


def loadData(filename):
    # 设置 API key 和 API base URL
    api_key = ""
    base_url = "https://api.132999.xyz/v1"
    all_excel_data = []
    client = OpenAI(
        api_key=api_key,
        base_url=base_url
    )

    # 打开xlsx文件
    workbook = openpyxl.load_workbook(filename)
    # 获取第一个工作表
    sheet = workbook.worksheets[0]
    # 读取数据
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        text_title = row[0].value
        text_chapter = row[1].value
        text_content = row[2].value
        prompt = text_title + " " + text_chapter + " "+ text_content+"\n根据上文构建一个恰当的问答对。"
        #print(prompt)
        chat_completion = client.chat.completions.create(
            messages=[
                {
                    "role": "user",
                    "content": prompt,
                }
            ],
            model="gpt-4-1106-preview",
        )
        text_result = chat_completion.choices[0].message.content
        data = {
            "title":text_title,
            "chapter":text_chapter,
            "content":text_content,
            "result":text_result,
        }
        data = dict(data)
        all_excel_data.append(data)
        df = pd.DataFrame(all_excel_data)
        excel_file_path = ""
        df.to_excel(excel_file_path, index=False)

if __name__ == '__main__':

    filename1 = ''
    selected_sentences = loadData(filename1)